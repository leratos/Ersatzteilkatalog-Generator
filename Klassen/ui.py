# -*- coding: utf-8 -*-
"""
Dieses Modul definiert das Hauptfenster der Anwendung (MainWindow).

Es ist verantwortlich für die Orchestrierung der Benutzeroberfläche,
das Laden der Projektdaten und die Interaktion zwischen den verschiedenen
Komponenten wie dem Konfigurations-Editor und dem Katalog-Generator.
"""

import json
import os
import shutil
import sys
from functools import partial

from PySide6 import QtCore, QtGui, QtWidgets

from Klassen.config import ConfigManager
from Klassen.editor_ui import ConfigEditorWindow
from Klassen.generator import DocxGenerator
from Klassen.stueckliste import BomProcessor


class MainWindow(QtWidgets.QMainWindow):
    """Das Hauptfenster der Anwendung."""

    def __init__(self, project_path: str):
        super().__init__()
        self.project_path = project_path
        self.config = ConfigManager(project_path)
        self.all_boms = {}
        self.item_lookup = {}
        self.output_column_configs = []
        self.grafik_column_index = -1
        self.is_dirty = False
        self.current_save_path = None

        self.setWindowTitle(
            f"Ersatzteilkatalog-Generator - Projekt: {os.path.basename(self.project_path)}"
        )
        self.resize(1200, 800)
        self._setup_ui()
        self._connect_signals()
        self._initialize_project()

    # --------------------------------------------------------------------------
    # --- UI Setup und Initialisierung ---
    # --------------------------------------------------------------------------

    def _setup_ui(self):
        """Erstellt alle UI-Elemente und ordnet sie im Layout an."""
        self.assembly_selector = QtWidgets.QComboBox()
        self.author_input = QtWidgets.QLineEdit()
        self.author_input.setPlaceholderText("Ihr Name")
        self.cover_graphic_button = QtWidgets.QPushButton("Deckblatt-Grafik...")
        self.tree_widget = QtWidgets.QTreeWidget()
        self.tree_widget.setSelectionMode(
            QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection
        )

        self.generate_button = QtWidgets.QPushButton("Katalog generieren")
        self.update_fields_checkbox = QtWidgets.QCheckBox(
            "Felder auto. aktualisieren (benötigt MS Word)"
        )
        if sys.platform == "win32":
            self.update_fields_checkbox.setChecked(True)
        else:
            self.update_fields_checkbox.setVisible(False)
            self.update_fields_checkbox.setChecked(False)
        self.save_button = QtWidgets.QPushButton("Auswahl speichern")
        self.load_button = QtWidgets.QPushButton("Auswahl laden")
        self.info_button = QtWidgets.QPushButton("Info / Copyright")
        self.config_button = QtWidgets.QPushButton("Konfiguration-Editor")

        top_layout = QtWidgets.QHBoxLayout()
        top_layout.addWidget(QtWidgets.QLabel("Hauptbaugruppe:"))
        top_layout.addWidget(self.assembly_selector, 2)
        top_layout.addWidget(QtWidgets.QLabel("Ersteller:"))
        top_layout.addWidget(self.author_input, 1)
        top_layout.addWidget(self.cover_graphic_button)

        bottom_layout = QtWidgets.QHBoxLayout()
        bottom_layout.addWidget(self.info_button)
        bottom_layout.addWidget(self.config_button)
        bottom_layout.addWidget(self.load_button)
        bottom_layout.addWidget(self.save_button)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.update_fields_checkbox)
        bottom_layout.addWidget(self.generate_button)

        main_layout = QtWidgets.QVBoxLayout()
        main_layout.addLayout(top_layout)
        main_layout.addWidget(self.tree_widget)
        main_layout.addLayout(bottom_layout)
        central_widget = QtWidgets.QWidget()
        central_widget.setLayout(main_layout)
        self.setCentralWidget(central_widget)

    def _connect_signals(self):
        """Verbindet die Signale der UI-Elemente mit ihren Funktionen."""
        self.assembly_selector.currentTextChanged.connect(
            self._on_assembly_selected
        )
        self.generate_button.clicked.connect(self._on_generate_button_clicked)
        self.save_button.clicked.connect(self._on_save_selection_clicked)
        self.load_button.clicked.connect(self._on_load_selection_clicked)
        self.cover_graphic_button.clicked.connect(
            self._on_assign_cover_graphic_clicked
        )
        self.tree_widget.itemChanged.connect(self._handle_item_changed)
        self.info_button.clicked.connect(self._show_info_dialog)
        self.config_button.clicked.connect(self._open_config_editor)

    def _initialize_project(self):
        """Prüft, ob ein Projekt existiert oder neu angelegt werden soll."""
        boms_path = os.path.join(self.project_path, "stücklisten")
        if not os.path.isdir(boms_path):
            reply = QtWidgets.QMessageBox.question(
                self,
                "Neues Projekt",
                "Der ausgewählte Ordner scheint kein Projekt zu sein. Möchten Sie hier ein neues Projekt erstellen?",
                QtWidgets.QMessageBox.StandardButton.Yes
                | QtWidgets.QMessageBox.StandardButton.No,
                QtWidgets.QMessageBox.StandardButton.Yes,
            )
            if reply == QtWidgets.QMessageBox.StandardButton.Yes:
                self._setup_new_project(boms_path)
            else:
                self.close()
                return
        else:
            self._load_project_data()
            self._auto_load_save_file()

    def _setup_new_project(self, boms_path):
        """Erstellt die Ordnerstruktur und initialen Dateien für ein neues Projekt."""
        try:
            os.makedirs(boms_path, exist_ok=True)
            os.makedirs(os.path.join(self.project_path, "Grafik"), exist_ok=True)
            
            master_template_folder = "Vorlagen"
            master_template_path = os.path.join(master_template_folder, "DOK-Vorlage.docm")
            if not os.path.exists(master_template_path):
                master_template_path = os.path.join(master_template_folder, "DOK-Vorlage.docx")
            if not os.path.exists(master_template_path):
                QtWidgets.QMessageBox.critical(self, "Fehler", "Keine Master-Vorlage im Ordner 'Vorlagen' gefunden.")
                return

            shutil.copy(master_template_path, self.project_path)
            self._prompt_for_boms(boms_path)
            self._load_project_data()
            self._on_save_selection_clicked()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler beim Erstellen des Projekts", str(e))

    # --------------------------------------------------------------------------
    # --- Kernlogik: Daten laden, verarbeiten und anzeigen ---
    # --------------------------------------------------------------------------

    def _load_project_data(self):
        """Lädt alle Projektdaten, wendet Regeln an und aktualisiert die UI."""
        boms_path = os.path.join(self.project_path, "stücklisten")
        if not os.path.isdir(boms_path):
            return

        self.config = ConfigManager(self.project_path)
        self._update_tree_columns()

        processor = BomProcessor(folder_path=boms_path, config_manager=self.config)
        self.all_boms = processor.run()
        self.load_data_into_ui(self.all_boms)

    def load_data_into_ui(self, all_boms: dict):
        """Befüllt die UI-Elemente mit den geladenen Daten."""
        self.all_boms = all_boms
        self.assembly_selector.blockSignals(True)
        self.assembly_selector.clear()
        if not self.all_boms:
            self.assembly_selector.addItem("Keine Stücklisten gefunden")
            self.assembly_selector.setEnabled(False)
            self.tree_widget.clear()
        else:
            self.assembly_selector.addItems(sorted(self.all_boms.keys()))
            self.assembly_selector.setEnabled(True)
        self.assembly_selector.blockSignals(False)
        if self.assembly_selector.count() > 0:
            self._on_assembly_selected(self.assembly_selector.currentText())

    def _on_assembly_selected(self, bom_znr: str):
        """Wird aufgerufen, wenn eine neue Hauptbaugruppe ausgewählt wird."""
        if bom_znr in self.all_boms:
            self._populate_tree(self.all_boms[bom_znr])

    def _update_tree_columns(self):
        """Liest die Konfiguration und passt die Spalten des Tree-Widgets an."""
        self.tree_widget.blockSignals(True)
        
        self.output_column_configs = self.config.config.get("output_columns", [])
        headers = [col_config.get("header", "") for col_config in self.output_column_configs]
        headers.append("Grafik")

        self.grafik_column_index = len(self.output_column_configs)
        self.tree_widget.setColumnCount(len(headers))
        self.tree_widget.setHeaderLabels(headers)

        header = self.tree_widget.header()
        for i, col_config in enumerate(self.output_column_configs):
            resize_mode = (
                QtWidgets.QHeaderView.ResizeMode.Stretch
                if col_config.get("id") == "std_benennung"
                else QtWidgets.QHeaderView.ResizeMode.Interactive
            )
            header.setSectionResizeMode(i, resize_mode)
            self.tree_widget.setColumnWidth(i, col_config.get("width_percent", 10) * 4)
        
        if header.count() > 0: 
            header.resizeSection(0, 50)
        self.tree_widget.blockSignals(False)

    def _populate_tree(self, main_assembly):
        """Baut den Baum für die ausgewählte Hauptbaugruppe neu auf."""
        self.tree_widget.blockSignals(True)
        self.tree_widget.clear()
        self.item_lookup.clear()
        
        root_item = QtWidgets.QTreeWidgetItem(self.tree_widget)
        root_item.setFlags(root_item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
        root_item.setCheckState(0, QtCore.Qt.CheckState.Checked)
        root_item.setFont(0, self._get_bold_font())

        benennung_col_index = next((i for i, c in enumerate(self.output_column_configs) if c.get("id") == "std_benennung"), 0)
        root_item.setText(benennung_col_index, f"{main_assembly.titel} ({main_assembly.zeichnungsnummer})")
        
        unique_id = str(main_assembly.zeichnungsnummer)
        root_item_data = {
            'Benennung': main_assembly.titel, 'Benennung_Formatiert': main_assembly.titel,
            'Teilenummer': main_assembly.zeichnungsnummer, 'is_assembly': True, 'unique_id': unique_id
        }
        root_item.setData(0, QtCore.Qt.ItemDataRole.UserRole, root_item_data)
        self.item_lookup[unique_id] = root_item

        if self.grafik_column_index != -1:
            assign_button = QtWidgets.QPushButton("Zuordnen...")
            assign_button.clicked.connect(partial(self._on_assign_graphic_clicked, root_item))
            self.tree_widget.setItemWidget(root_item, self.grafik_column_index, assign_button)

        self._add_children_recursively(root_item, main_assembly)
        self.tree_widget.expandAll()
        self.tree_widget.blockSignals(False)

    def _add_children_recursively(self, parent_item, bom_obj):
        """Fügt Kind-Elemente rekursiv hinzu."""
        sorted_positions = sorted(bom_obj.positionen, key=lambda p: float(p.get('POS', 'inf')))
        for position in sorted_positions:
            child_item = QtWidgets.QTreeWidgetItem(parent_item)
            child_item.setFlags(child_item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
            child_item.setCheckState(0, QtCore.Qt.CheckState.Checked)

            for i, col_config in enumerate(self.output_column_configs):
                source_id = col_config.get("source_id") or col_config.get("id")
                cell_text = position.get(source_id, "")
                if source_id == 'POS' and isinstance(cell_text, (int, float)):
                    cell_text = f"{cell_text:g}"
                child_item.setText(i, str(cell_text))
                if not col_config.get("source_id"):
                    child_item.setFlags(child_item.flags() | QtCore.Qt.ItemFlag.ItemIsEditable)
            
            is_assembly = 'sub_assembly' in position
            if is_assembly:
                child_item.setFont(0, self._get_bold_font())
                if self.grafik_column_index != -1:
                    assign_button = QtWidgets.QPushButton("Zuordnen...")
                    assign_button.clicked.connect(partial(self._on_assign_graphic_clicked, child_item))
                    self.tree_widget.setItemWidget(child_item, self.grafik_column_index, assign_button)

            unique_id = f"{bom_obj.zeichnungsnummer}_{position.get('POS', '')}"
            position_data = position.copy()
            position_data['is_assembly'] = is_assembly
            position_data['unique_id'] = unique_id
            child_item.setData(0, QtCore.Qt.ItemDataRole.UserRole, position_data)
            self.item_lookup[unique_id] = child_item
            
            if is_assembly: self._add_children_recursively(child_item, position['sub_assembly'])

    # --------------------------------------------------------------------------
    # --- Event Handler und Slots ---
    # --------------------------------------------------------------------------

    def _open_config_editor(self):
        """Öffnet den zentralen Konfigurations-Editor."""
        main_bom_znr = self.assembly_selector.currentText()
        main_bom_obj = self.all_boms.get(main_bom_znr)
        excel_columns = []
        if main_bom_obj:
             try:
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter
                workbook = load_workbook(main_bom_obj.filepath, read_only=True, data_only=True)
                sheet = workbook['Import']
                for cell in sheet[5]:
                    if cell.value:
                        excel_columns.append(f"{get_column_letter(cell.column)} - {cell.value}")
             except Exception as e:
                print(f"Fehler beim Lesen der Muster-Header für den Editor: {e}")

        editor_dialog = ConfigEditorWindow(self.config, excel_columns, self)
        
        if editor_dialog.exec():
            print("INFO: Konfiguration wurde gespeichert. Lade Projekt neu...")
            self._load_project_data()
            QtWidgets.QMessageBox.information(self, "Konfiguration gespeichert", "Die Konfiguration wurde aktualisiert. Das Projekt wird neu geladen, um alle Änderungen anzuwenden.")

    def _on_generate_button_clicked(self):
        """Startet den Prozess zur Erstellung des Word-Dokuments."""
        root = self.tree_widget.invisibleRootItem()
        if root.childCount() == 0: return
        root_item = root.child(0)
        if not root_item or root_item.checkState(0) == QtCore.Qt.CheckState.Unchecked: return
        
        hierarchical_data = self._collect_hierarchical_data(root_item)
        if not hierarchical_data or (not hierarchical_data.get('children') and not hierarchical_data.get('is_assembly')): return
        
        main_bom_znr = self.assembly_selector.currentText()
        main_bom_obj = self.all_boms.get(main_bom_znr)
        if not main_bom_obj: return
        
        # --- KORRIGIERT: Dynamische Pfad- und Filterlogik für .docm ---
        template_path_docm = os.path.join(self.project_path, "DOK-Vorlage.docm")
        template_path_docx = os.path.join(self.project_path, "DOK-Vorlage.docx")
        
        template_path, file_filter, suggested_ext = "", "", ""
        if os.path.exists(template_path_docm):
            template_path = template_path_docm
            file_filter = "Word-Dokument mit Makros (*.docm)"
            suggested_ext = ".docm"
        elif os.path.exists(template_path_docx):
            template_path = template_path_docx
            file_filter = "Word-Dokument (*.docx)"
            suggested_ext = ".docx"
        else:
            QtWidgets.QMessageBox.critical(self, "Fehler", "Keine DOK-Vorlage (.docm oder .docx) im Projektordner gefunden.")
            return
            
        suggested_filename = f"Ersatzteilkatalog_{main_bom_znr}{suggested_ext}"
        output_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Katalog speichern",
            os.path.join(self.project_path, suggested_filename),
            f"{file_filter};;Alle Dateien (*.*)"
        )
        # --- ENDE DER KORREKTUR ---
        
        if not output_path: return
        
        generator = DocxGenerator(
            data=hierarchical_data, main_bom=main_bom_obj, author_name=self.author_input.text(),
            template_path=template_path, output_path=output_path,
            auto_update_fields=self.update_fields_checkbox.isChecked(),
            project_path=self.project_path, config_manager=self.config
        )
        if generator.run():
            self._show_generation_success_dialog(output_path)
        else:
            QtWidgets.QMessageBox.critical(self, "Fehler", "Beim Erstellen des Katalogs ist ein Fehler aufgetreten.")

    def _handle_item_changed(self, item, column):
        """Behandelt Änderungen an Items (Checkbox oder manuelle Eingabe)."""
        self.is_dirty = True
        self.tree_widget.blockSignals(True)
        try:
            if column == 0:
                self._set_children_checkstate(item, item.checkState(0))
            elif column > 0 and (item.flags() & QtCore.Qt.ItemFlag.ItemIsEditable):
                col_config = self.output_column_configs[column]
                if not col_config.get("source_id"):
                    data = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
                    if data:
                        data[col_config.get("id")] = item.text(column)
                        item.setData(0, QtCore.Qt.ItemDataRole.UserRole, data)
        finally:
            self.tree_widget.blockSignals(False)

    # --------------------------------------------------------------------------
    # --- Speichern und Laden der Auswahl ---
    # --------------------------------------------------------------------------

    def _auto_load_save_file(self):
        """Sucht und lädt automatisch eine Projekt-Speicherdatei."""
        for filename in os.listdir(self.project_path):
            if filename.lower().startswith("projekt_") and filename.lower().endswith(".json"):
                self._load_selection_from_file(os.path.join(self.project_path, filename))
                return

    def _on_load_selection_clicked(self):
        load_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Auswahl laden", self.project_path, "JSON-Dateien (*.json)")
        if load_path: self._load_selection_from_file(load_path)

    def _load_selection_from_file(self, file_path):
        """Lädt eine Speicherdatei und wendet die Einstellungen an."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                load_data = json.load(f)
            
            self.current_save_path = file_path
            self.author_input.setText(load_data.get('author', ''))
            main_bom_znr = load_data.get('main_bom_znr')
            
            if main_bom_znr in self.all_boms:
                self.assembly_selector.setCurrentText(main_bom_znr)
                manual_data_to_load = load_data.get('manual_data', {})
                if manual_data_to_load:
                     QtCore.QTimer.singleShot(150, lambda: self._apply_manual_data(manual_data_to_load))
                QtCore.QTimer.singleShot(100, lambda: self._apply_loaded_selection(load_data.get('unchecked_item_ids', [])))
            else:
                QtWidgets.QMessageBox.warning(self, "Warnung", f"Die in '{os.path.basename(file_path)}' gespeicherte Hauptbaugruppe '{main_bom_znr}' wurde nicht gefunden.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler beim Laden", f"Die Speicherdatei konnte nicht geladen werden:\n{e}")
        finally:
            self.is_dirty = False

    def _on_save_selection_clicked(self):
        """Speichert die aktuelle Auswahl in eine JSON-Datei."""
        if not self.assembly_selector.currentText(): return
        
        save_path = self.current_save_path
        if not save_path:
            suggested_filename = f"projekt_{os.path.basename(self.project_path)}.json"
            path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Auswahl speichern unter", os.path.join(self.project_path, suggested_filename), "JSON-Dateien (*.json)")
            if not path: return
            save_path = path

        save_data = {
            'main_bom_znr': self.assembly_selector.currentText(),
            'author': self.author_input.text(),
            'unchecked_item_ids': self._collect_unchecked_items(),
            'manual_data': self._collect_manual_data()
        }
        try:
            with open(save_path, 'w', encoding='utf-8') as f:
                json.dump(save_data, f, indent=4, ensure_ascii=False)
            self.current_save_path = save_path
            self.is_dirty = False
            print(f"INFO: Auswahl erfolgreich in '{os.path.basename(save_path)}' gespeichert.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler beim Speichern", str(e))

    # --------------------------------------------------------------------------
    # --- Hilfsmethoden ---
    # --------------------------------------------------------------------------

    def _collect_unchecked_items(self) -> list:
        """Sammelt die unique_ids aller abgewählten Items."""
        collection = []
        iterator = QtWidgets.QTreeWidgetItemIterator(self.tree_widget)
        while iterator.value():
            item = iterator.value()
            if item.checkState(0) == QtCore.Qt.CheckState.Unchecked:
                item_data = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
                if item_data and item_data.get('unique_id'):
                    collection.append(item_data.get('unique_id'))
            iterator += 1
        return collection
        
    def _collect_manual_data(self) -> dict:
        """Sammelt alle manuell eingegebenen Daten."""
        manual_data = {}
        iterator = QtWidgets.QTreeWidgetItemIterator(self.tree_widget)
        while iterator.value():
            item = iterator.value()
            item_data = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
            if not item_data:
                iterator += 1
                continue
            item_manual_values = {}
            for col_idx, col_config in enumerate(self.output_column_configs):
                if not col_config.get("source_id"):
                    manual_key = col_config.get("id")
                    if manual_key in item_data and item_data[manual_key]:
                        item_manual_values[manual_key] = item_data[manual_key]
            if item_manual_values:
                unique_id = item_data.get('unique_id')
                if unique_id: manual_data[unique_id] = item_manual_values
            iterator += 1
        return manual_data

    def _collect_hierarchical_data(self, parent_item):
        """Sammelt die Daten aus dem Baum in einer hierarchischen Struktur."""
        if parent_item is None or parent_item.checkState(0) == QtCore.Qt.CheckState.Unchecked: return None
        parent_data = parent_item.data(0, QtCore.Qt.ItemDataRole.UserRole)
        if not parent_data: return None
        
        data_copy = parent_data.copy()
        data_copy.pop('sub_assembly', None)
        data_copy['children'] = []
        
        for i in range(parent_item.childCount()):
            child_item = parent_item.child(i)
            child_data = self._collect_hierarchical_data(child_item)
            if child_data: data_copy['children'].append(child_data)
        return data_copy

    def _apply_loaded_selection(self, unchecked_id_list: list):
        """Wendet die Checkbox-Zustände aus einer geladenen Datei an."""
        self.tree_widget.blockSignals(True)
        for unique_id in unchecked_id_list:
            if unique_id in self.item_lookup:
                self.item_lookup[unique_id].setCheckState(0, QtCore.Qt.CheckState.Unchecked)
        self.tree_widget.blockSignals(False)

    def _apply_manual_data(self, manual_data: dict):
        """Wendet manuelle Daten aus einer geladenen Datei an."""
        if not manual_data: return
        self.tree_widget.blockSignals(True)
        try:
            for unique_id, values_to_set in manual_data.items():
                if unique_id in self.item_lookup:
                    item = self.item_lookup[unique_id]
                    item_data = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
                    if item_data:
                        item_data.update(values_to_set)
                        item.setData(0, QtCore.Qt.ItemDataRole.UserRole, item_data)
                    for key, value in values_to_set.items():
                        for col_idx, col_config in enumerate(self.output_column_configs):
                            if col_config.get("id") == key:
                                item.setText(col_idx, str(value))
                                break
        finally:
            self.tree_widget.blockSignals(False)

    def _set_children_checkstate(self, parent_item, state):
        """Setzt den Check-Status aller Kind-Elemente rekursiv."""
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            child.setCheckState(0, state)
            if child.childCount() > 0:
                self._set_children_checkstate(child, state)

    def _on_assign_cover_graphic_clicked(self):
        """Öffnet einen Dialog zur Auswahl der Deckblatt-Grafik."""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Deckblatt-Grafik auswählen", self.project_path, "Bilder (*.png *.jpg *.jpeg)")
        if file_path:
            grafik_folder = os.path.join(self.project_path, "Grafik")
            os.makedirs(grafik_folder, exist_ok=True)
            _, file_extension = os.path.splitext(file_path)
            destination_path = os.path.join(grafik_folder, f"EL{file_extension}")
            try:
                shutil.copy(file_path, destination_path)
                QtWidgets.QMessageBox.information(self, "Erfolg", "Deckblatt-Grafik wurde erfolgreich gespeichert.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Fehler beim Kopieren", f"Die Grafik konnte nicht gespeichert werden:\n{e}")

    def _on_assign_graphic_clicked(self, item):
        """Öffnet einen Dialog zur Auswahl einer Baugruppen-Grafik."""
        item_data = item.data(0, QtCore.Qt.ItemDataRole.UserRole)
        if not item_data: return
        zeichnung_nr = item_data.get('Teilenummer')
        if not zeichnung_nr:
            QtWidgets.QMessageBox.warning(self, "Fehler", "Für diese Position wurde keine Zeichnungsnummer gefunden.")
            return
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Grafik auswählen", self.project_path, "Bilder (*.png *.jpg *.jpeg)")
        if file_path:
            grafik_folder = os.path.join(self.project_path, "Grafik")
            os.makedirs(grafik_folder, exist_ok=True)
            _, file_extension = os.path.splitext(file_path)
            destination_path = os.path.join(grafik_folder, f"{zeichnung_nr}{file_extension}")
            try:
                shutil.copy(file_path, destination_path)
                QtWidgets.QMessageBox.information(self, "Erfolg", f"Grafik wurde erfolgreich für {zeichnung_nr} gespeichert.")
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Fehler beim Kopieren", f"Die Grafik konnte nicht gespeichert werden:\n{e}")

    def _show_generation_success_dialog(self, output_path: str):
        """Zeigt einen Erfolgsdialog nach der Katalogerstellung an."""
        msg_box = QtWidgets.QMessageBox()
        msg_box.setIcon(QtWidgets.QMessageBox.Icon.Information)
        msg_box.setText("Erfolg!")
        info_text = f"Der Katalog wurde erfolgreich gespeichert."
        if not self.update_fields_checkbox.isChecked() and sys.platform == "win32":
            info_text += "\n\nWICHTIG: Um den Schriftkopf und das Inhaltsverzeichnis zu aktualisieren, öffnen Sie das Dokument, drücken Sie Strg+A und dann F9."
        msg_box.setInformativeText(info_text)
        msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok | QtWidgets.QMessageBox.StandardButton.Open)
        button_open = msg_box.button(QtWidgets.QMessageBox.StandardButton.Open)
        button_open.setText("Datei öffnen")
        if msg_box.exec() == QtWidgets.QMessageBox.StandardButton.Open:
            try:
                os.startfile(output_path)
            except AttributeError:
                print(f"Datei kann nicht automatisch geöffnet werden. Pfad: {output_path}")

    def _show_info_dialog(self):
        """Zeigt ein Fenster mit Copyright-Informationen an."""
        msg_box = QtWidgets.QMessageBox(self)
        msg_box.setWindowTitle("Information und Copyright")
        msg_box.setIcon(QtWidgets.QMessageBox.Icon.Information)
        msg_box.setText(
            "<b>Ersatzteilkatalog-Generator</b><br><br>"
            "Alle Rechte für diese Software liegen bei:<br>"
            "<b>Marcus Kohtz (Signz-vision.de)</b>"
        )
        msg_box.setInformativeText(
            "Eine Weitergabe, Vervielfältigung oder Nutzung dieser Anwendung, "
            "ganz oder in Teilen, ist ohne die ausdrückliche schriftliche "
            "Genehmigung des Urhebers nicht gestattet."
        )
        msg_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def _get_bold_font(self):
        """Gibt eine fette Schriftart zurück."""
        font = QtGui.QFont()
        font.setBold(True)
        return font
    
    def _prompt_for_boms(self, boms_path):
        """Fragt den Benutzer nach Stücklisten für ein neues Projekt."""
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(self, "Wählen Sie Stücklisten für das neue Projekt", "", "Excel-Dateien (*.xlsm *.xlsx)")
        if not files:
            QtWidgets.QMessageBox.warning(self, "Abbruch", "Keine Stücklisten ausgewählt. Das Projekt ist leer.")
            return
        for file_path in files:
            shutil.copy(file_path, boms_path)

    def closeEvent(self, event):
        """Wird aufgerufen, wenn das Fenster geschlossen wird."""
        if self.is_dirty:
            reply = QtWidgets.QMessageBox.question(
                self, 'Ungespeicherte Änderungen',
                "Es gibt ungespeicherte Änderungen. Möchten Sie sie vor dem Schließen speichern?",
                QtWidgets.QMessageBox.StandardButton.Save |
                QtWidgets.QMessageBox.StandardButton.Discard |
                QtWidgets.QMessageBox.StandardButton.Cancel
            )
            if reply == QtWidgets.QMessageBox.StandardButton.Save:
                self._on_save_selection_clicked()
                if not self.is_dirty: event.accept()
                else: event.ignore()
            elif reply == QtWidgets.QMessageBox.StandardButton.Discard:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()
