# -*- coding: utf-8 -*-
"""
Dieses Modul definiert das Konfigurationsfenster (Editor).

Es ermöglicht dem Benutzer, die Zuordnung von Excel-Spalten zu den
internen Datenfeldern des Programms dynamisch zu ändern, ohne den
Quellcode bearbeiten zu müssen.
"""

from PySide6 import QtWidgets, QtCore
import openpyxl
from openpyxl.utils import get_column_letter
import re
import time

class ConfigEditorWindow(QtWidgets.QDialog):
    """
    Ein Dialogfenster zur Bearbeitung der Projekt-Konfiguration (mapping.json).
    """
    def __init__(self, config_manager, available_excel_columns, parent=None):
        """
        Initialisiert das Editor-Fenster.

        Args:
            config_manager (ConfigManager): Die Instanz des ConfigManagers,
                                            die die Konfiguration verwaltet.
            parent (QWidget, optional): Das übergeordnete Widget. Defaults to None.
        """
        super().__init__(parent)
        self.config_manager = config_manager
        self.available_columns = [""] + available_excel_columns
        self.available_excel_columns = available_excel_columns
        self.setWindowTitle("Katalog-Editor")
        self.setMinimumSize(800, 600)

        self.layout = QtWidgets.QVBoxLayout(self)
        self.tabs = QtWidgets.QTabWidget()
        self.layout.addWidget(self.tabs)
        
        self.mapping_tab = QtWidgets.QWidget()
        self.layout_tab = QtWidgets.QWidget()

        self.tabs.addTab(self.mapping_tab, "Spaltenzuordnung (Import)")
        self.tabs.addTab(self.layout_tab, "Katalog-Layout (Export)")

        self._setup_mapping_tab()
        self._setup_layout_tab()

        self.button_box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.StandardButton.Save | 
            QtWidgets.QDialogButtonBox.StandardButton.Cancel
        )
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)

    def _setup_layout_tab(self):
        """Erstellt die UI für den Katalog-Layout-Editor."""
        layout = QtWidgets.QVBoxLayout(self.layout_tab)
        
        self.layout_table = QtWidgets.QTableWidget()
        self.layout_table.setColumnCount(3)
        self.layout_table.setHorizontalHeaderLabels(
            ["Spaltenüberschrift", "Datenquelle (interne ID)", "Breite (%)"] # Geändert
        )
        self.layout_table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.layout_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.layout_table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        
        self._populate_layout_table()

        button_layout = QtWidgets.QHBoxLayout()
        add_button = QtWidgets.QPushButton("Benutzerdefinierte Spalte hinzufügen")
        remove_button = QtWidgets.QPushButton("Ausgewählte Spalte entfernen")
        up_button = QtWidgets.QPushButton("Hoch")
        down_button = QtWidgets.QPushButton("Runter")
        
        button_layout.addWidget(add_button)
        button_layout.addWidget(remove_button)
        button_layout.addStretch()
        button_layout.addWidget(up_button)
        button_layout.addWidget(down_button)
        
        layout.addWidget(self.layout_table)
        layout.addLayout(button_layout)
        
        add_button.clicked.connect(self._add_layout_row)
        remove_button.clicked.connect(self._remove_layout_row)
        up_button.clicked.connect(self._move_row_up) 
        down_button.clicked.connect(self._move_row_down)

    def _create_row_widgets(self, row, data, available_ids):
        """Erstellt die Widgets für eine einzelne Zeile der Tabelle."""
        is_standard = data.get("type") == "standard"
        
        item_header = QtWidgets.QTableWidgetItem(data.get("header", "Neue Spalte"))
        self.layout_table.setItem(row, 0, item_header)
        item_header.setData(QtCore.Qt.ItemDataRole.UserRole, data)

        source_id = data.get("source_id", "")
        if is_standard:
            label = QtWidgets.QLabel(f"<i>{source_id} (Standard)</i>")
            label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            self.layout_table.setCellWidget(row, 1, label)
        else:
            combo_id = QtWidgets.QComboBox()
            combo_id.addItems(available_ids)
            combo_id.setCurrentText(source_id)
            self.layout_table.setCellWidget(row, 1, combo_id)

        spin_width = QtWidgets.QSpinBox()
        spin_width.setRange(1, 100)
        spin_width.setSuffix(" %")
        spin_width.setValue(int(data.get("width_percent", 10)))
        self.layout_table.setCellWidget(row, 2, spin_width)
        

    def _get_row_data(self, row):
        """Liest alle Daten aus einer Zeile der Layout-Tabelle."""
        header_item = self.layout_table.item(row, 0)
        if not header_item: return None
        
        stored_data = header_item.data(QtCore.Qt.ItemDataRole.UserRole)
        if not isinstance(stored_data, dict):
            # Fallback für den unwahrscheinlichen Fall, dass die Daten veraltet sind
            stored_data = {"id": None, "type": "custom"}

        col_type = stored_data.get("type")
        col_id = stored_data.get("id") # Wichtig: Die ursprüngliche ID wird hier geholt.
        
        source_widget = self.layout_table.cellWidget(row, 1)
        source_id = ""
        if isinstance(source_widget, QtWidgets.QComboBox):
            source_id = source_widget.currentText()
        elif isinstance(source_widget, QtWidgets.QLabel):
            match = re.search(r'<i>(.*?) \(Standard\)</i>', source_widget.text())
            if match: source_id = match.group(1)

        width_widget = self.layout_table.cellWidget(row, 2)
        
        return {
            "id": col_id, # Wichtig: Die ID wird in die neue Konfiguration übernommen.
            "header": header_item.text(),
            "width_percent": width_widget.value(),
            "source_id": source_id,
            "type": col_type
        }

    def _setup_mapping_tab(self):
        """Erstellt die Widgets für den "Spaltenzuordnung"-Tab."""
        layout = QtWidgets.QVBoxLayout(self.mapping_tab)

        header_group = QtWidgets.QGroupBox("Header-Felder (Zellen)")
        header_layout = QtWidgets.QGridLayout()
        header_group.setLayout(header_layout)
        
        self.header_widgets = {}
        header_config = self.config_manager.config.get("header_mapping", {})
        for row, (key, value) in enumerate(header_config.items()):
            label = QtWidgets.QLabel(f"{key}:")
            line_edit = QtWidgets.QLineEdit(value)
            self.header_widgets[key] = line_edit
            header_layout.addWidget(label, row, 0)
            header_layout.addWidget(line_edit, row, 1)

        self.column_group = QtWidgets.QGroupBox("Positions-Felder (Spalten)")
        self.column_layout = QtWidgets.QGridLayout()
        self.column_group.setLayout(self.column_layout)

        self.column_widgets = {}
        
        self._create_column_comboboxes()
            
        layout.addWidget(header_group)
        layout.addWidget(self.column_group)
        layout.addStretch()

    
    def _add_column(self):
        """Fügt ein ausgewähltes Feld zur Liste der Katalog-Spalten hinzu."""
        selected_item = self.available_fields_list.currentItem()
        if selected_item:
            # Füge den Header-Text zur UI-Liste hinzu.
            self.selected_columns_list.addItem(selected_item.text())

    def _create_column_comboboxes(self):
        """(Neu) Erstellt oder aktualisiert die Dropdown-Menüs für die Spalten."""
        while self.column_layout.count():
            child = self.column_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        self.column_widgets = {}
        column_config = self.config_manager.config.get("column_mapping", {})
        
        for row, (key, current_value) in enumerate(column_config.items()):
            label = QtWidgets.QLabel(f"{key}:")
            combo_box = QtWidgets.QComboBox()
            combo_box.addItems(self.available_columns)
            
            full_text_to_set = ""
            for option in self.available_columns:
                if option.startswith(current_value + ' -'):
                    full_text_to_set = option
                    break
            
            if full_text_to_set:
                combo_box.setCurrentText(full_text_to_set)
            else:
                # Fallback, falls der Buchstabe nicht in den Optionen ist
                combo_box.addItem(current_value)
                combo_box.setCurrentText(current_value)

            self.column_widgets[key] = combo_box
            self.column_layout.addWidget(label, row, 0)
            self.column_layout.addWidget(combo_box, row, 1)

    def _load_sample_bom(self):
        """
        Öffnet eine Excel-Datei, liest die Spaltenüberschriften aus,
        erweitert die Konfiguration um neue Felder und aktualisiert die komplette UI.
        """
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Muster-Stückliste auswählen", "", "Excel-Dateien (*.xlsm *.xlsx)")

        if not file_path:
            return
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook['Import']
            
            headers_full_str = []  # z.B. ["Q - Änderungsindex"]
            header_map = {}        # z.B. {"Änderungsindex": "Q"}

            for cell in sheet[5]:  # Annahme: Header sind in Zeile 5
                if cell.value:
                    col_letter = get_column_letter(cell.column)
                    header_name = str(cell.value).strip()
                    
                    if header_name:
                        headers_full_str.append(f"{col_letter} - {header_name}")
                        header_map[header_name] = col_letter

            if headers_full_str:
                # --- KERN DER NEUEN LOGIK ---
                # 1. Hole die aktuelle Konfiguration aus dem Manager
                new_config = self.config_manager.config
                current_mapping = new_config.get("column_mapping", {})
                
                # 2. Füge neue, bisher unbekannte Felder zur Konfiguration hinzu
                new_fields_added = False
                for name, letter in header_map.items():
                    if name not in current_mapping:
                        current_mapping[name] = letter  # z.B. current_mapping["Änderungsindex"] = "Q"
                        new_fields_added = True

                # 3. Aktualisiere die Konfiguration im Speicher (wird erst bei "Save" geschrieben)
                new_config["column_mapping"] = current_mapping
                self.config_manager.config = new_config

                # 4. Aktualisiere die UI-Komponenten, die von der Konfiguration abhängen
                self.available_columns = [""] + headers_full_str
                
                # Baue beide Tabs komplett neu auf, da sich die Konfig geändert haben könnte
                self._create_column_comboboxes()  # Baut Import-Tab neu auf (zeigt jetzt "Änderungsindex")
                self._populate_layout_table()     # Baut Export-Tab neu auf (Dropdown hat jetzt "Änderungsindex")

                msg = (f"{len(headers_full_str)} Spalten geladen. Die Dropdown-Listen wurden aktualisiert.")
                if new_fields_added:
                    msg += "\nNeue Felder wurden im 'Import'-Tab ergänzt."
                
                QtWidgets.QMessageBox.information(self, "Erfolg", msg)

            else:
                QtWidgets.QMessageBox.warning(self, "Fehler", "Konnte keine Spaltenüberschriften in Zeile 5 finden.")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Fehler beim Lesen der Datei", str(e))
            
    def accept(self):
        """Wird aufgerufen, wenn der Benutzer auf "Speichern" klickt."""
        new_header_mapping = {}
        new_column_mapping = {}

        for key, widget in self.header_widgets.items():
            new_header_mapping[key] = widget.text().upper()
            
        new_column_mapping = {}
        for key, widget in self.column_widgets.items():
            # Lese den ausgewählten Text aus der ComboBox
            # und nehme nur den Spaltenbuchstaben am Anfang.
            selected_text = widget.currentText()
            new_column_mapping[key] = selected_text.split(' - ')[0]

        new_output_columns = []
        for row in range(self.layout_table.rowCount()):
            data = self._get_row_data(row)
            if data:
                new_output_columns.append(data)
            
        new_config = self.config_manager.config
        new_config["header_mapping"] = new_header_mapping
        new_config["column_mapping"] = new_column_mapping
        new_config["output_columns"] = new_output_columns
        self.config_manager.save_config(new_config)
        super().accept()

    def _populate_layout_table(self):
        """Füllt die Layout-Tabelle mit den Daten aus der Konfiguration."""
        output_config = self.config_manager.config.get("output_columns", [])
        available_ids = self._get_complete_source_ids()
        
        self.layout_table.setRowCount(len(output_config))
        for row_idx, col_data in enumerate(output_config):
            self._create_row_widgets(row_idx, col_data, available_ids)

    def _add_layout_row(self):
        """Fügt eine neue, leere benutzerdefinierte Zeile zur Layout-Tabelle hinzu."""
        row = self.layout_table.rowCount()
        self.layout_table.insertRow(row)
        available_ids = self._get_complete_source_ids()
        new_id = f"custom_{int(time.time() * 1000)}"
        new_col_data = {
            "id": new_id,
            "type": "custom",
            "header": "Neue Spalte",
            "width_percent": 15,
            "source_id": ""
        }
        self._create_row_widgets(row, new_col_data, available_ids)

    def _remove_layout_row(self):
        row = self.layout_table.currentRow()
        if row >= 0:
            item = self.layout_table.item(row, 0)
            if item:
                # ÄNDERUNG: Greife auf den 'type' im gespeicherten Dictionary zu.
                stored_data = item.data(QtCore.Qt.ItemDataRole.UserRole)
                if isinstance(stored_data, dict) and stored_data.get("type") == "custom":
                    self.layout_table.removeRow(row)
                else:
                    QtWidgets.QMessageBox.warning(self, "Fehler", "Standard-Spalten können nicht entfernt werden.")

    def _move_row_up(self):
        """Bewegt die ausgewählte Zeile eine Position nach oben."""
        row = self.layout_table.currentRow()
        if row > 0:
            data = self._get_row_data(row)
            self.layout_table.removeRow(row)
            self.layout_table.insertRow(row - 1)
            self._create_row_widgets(row - 1, data, self._get_complete_source_ids())
            self.layout_table.setCurrentCell(row - 1, 0)
            
    def _move_row_down(self):
        """Bewegt die ausgewählte Zeile eine Position nach unten."""
        row = self.layout_table.currentRow()
        if 0 <= row < self.layout_table.rowCount() - 1:
            data = self._get_row_data(row)
            self.layout_table.removeRow(row)
            self.layout_table.insertRow(row + 1)
            self._create_row_widgets(row + 1, data, self._get_complete_source_ids())
            self.layout_table.setCurrentCell(row + 1, 0)
    
    def _copy_row_content(self, from_row, to_row):
        # Hilfsfunktion zum Kopieren von Zeileninhalten
        data = { "header": self.layout_table.item(from_row, 0).text(), "type": self.layout_table.item(from_row, 0).data(QtCore.Qt.ItemDataRole.UserRole) }
        source_widget = self.layout_table.cellWidget(from_row, 1)
        if isinstance(source_widget, QtWidgets.QComboBox): data["source_id"] = source_widget.currentText()
        else: 
            match = re.search(r'<i>(.*?) \(Standard\)</i>', source_widget.text())
            if match: data["source_id"] = match.group(1)
        
        width_widget = self.layout_table.cellWidget(from_row, 2)
        if width_widget: data["width_percent"] = width_widget.value()
        
        self._create_row_widgets(to_row, data, self.config_manager.get_all_available_data_ids())

    def _get_complete_source_ids(self) -> list:
        """
        Erstellt eine vollständige Liste aller verfügbaren Daten-IDs durch die Kombination
        von Konfiguration, generierten Feldern und den Spalten aus der Muster-Stückliste.
        Dies ist die korrigierte Logik, um das Problem zu beheben.
        """
        # 1. Felder aus der bestehenden Konfiguration (z.B. POS, Teilenummer)
        input_fields_from_config = list(self.config_manager.config.get("column_mapping", {}).keys())

        # 2. Vom Programm generierte Felder
        generated_fields = [
            "Benennung_Formatiert", "Menge",
            "Bestellnummer_Kunde", "Information", "Seite"
        ]

        # 3. Felder, die aus den Spaltenüberschriften der Muster-Stückliste stammen
        # self.available_columns enthält Strings wie "A - POS", "R - Gewicht"
        fields_from_sample_bom = []
        for col_str in self.available_columns:
            if ' - ' in col_str:
                # Extrahiere den Namen nach dem Trennzeichen ' - '
                field_name = col_str.split(' - ', 1)[1]
                fields_from_sample_bom.append(field_name)

        # 4. Alle Quellen zusammenführen, Duplikate entfernen und sortieren
        all_fields = set(input_fields_from_config + generated_fields + fields_from_sample_bom)

        # Gib eine sortierte Liste mit einer leeren Start-Option zurück
        return [""] + sorted(list(all_fields))